import numpy as np
import pandas as pd
import plotly.graph_objects as go

from backend.services.report_export import (
    _append_plot_data_sheet,
    _board_export_sections,
    _normalize_plotly_trace,
    _split_table_payload,
    _strip_duplicate_note_heading,
    build_board_export,
)
from backend.tools.impl.plotly_tool import apply_default_chart_style


def test_strip_duplicate_note_heading_removes_suty_prefix() -> None:
    content = "Суть Лидером по вкладу является акция NREH."
    cleaned = _strip_duplicate_note_heading(content, "Суть")
    assert cleaned.startswith("Лидером")
    assert "Суть Суть" not in cleaned


def test_strip_duplicate_note_heading_removes_bold_line() -> None:
    content = "**Суть**\n\nПортфель демонстрирует концентрацию."
    cleaned = _strip_duplicate_note_heading(content, "**Суть**")
    assert cleaned.startswith("Портфель")
    assert "**Суть**" not in cleaned.splitlines()[0]


def test_board_export_sections_group_by_question() -> None:
    artifacts = [
        {"id": "plot1", "type": "plot", "text": "chart_a"},
        {"id": "note1", "type": "note", "text": "note"},
        {"id": "plot2", "type": "plot", "text": "chart_b"},
    ]
    sections = [
        {"label": "Вопрос 1: структура", "artifact_ids": ["plot1", "note1"]},
        {"label": "Вопрос 2: концентрация", "artifact_ids": ["plot2"]},
    ]
    grouped = _board_export_sections(artifacts, sections)
    assert [label for label, _ in grouped] == [
        "Вопрос 1: структура",
        "Вопрос 2: концентрация",
    ]
    assert grouped[0][1][0]["id"] == "plot1"
    assert grouped[1][1][0]["id"] == "plot2"


def test_board_bar_export_uses_single_color() -> None:
    from backend.services import report_export

    df = pd.DataFrame({"ticker": ["A", "B", "C"], "value": [1.0, 2.0, 3.0]})
    fig = apply_default_chart_style(go.Figure(go.Bar(x=df["ticker"], y=df["value"], orientation="v")))
    colors_before = fig.data[0].marker.color
    assert isinstance(colors_before, (list, tuple)) and len(colors_before) > 1

    report_export._apply_board_bar_colors(fig)
    assert fig.data[0].marker.color == report_export._BOARD_BAR_COLOR


def test_report_plotly_trace_decodes_typed_array_bar_points() -> None:
    fig = go.Figure(go.Bar(x=np.array([10.0, 20.0, 30.0]), y=["a", "b", "c"], orientation="h"))
    trace = fig.to_plotly_json()["data"][0]

    normalized = _normalize_plotly_trace(trace, 0)

    assert normalized["marker"]["color"] == ["#2563eb", "#7c3aed", "#0f766e"]


def test_report_xlsx_plot_data_decodes_typed_arrays() -> None:
    from openpyxl import Workbook

    fig = go.Figure(go.Bar(x=np.array([10.0, 20.0, 30.0]), y=["a", "b", "c"], name="series"))
    artifact = {
        "id": "plot1",
        "type": "plot",
        "text": "Chart",
        "data": {"format": "plotly-json", "data": fig.to_plotly_json()},
        "meta": {},
    }
    workbook = Workbook()
    summary_sheet = workbook.active

    _append_plot_data_sheet(
        workbook=workbook,
        artifact=artifact,
        artifact_index={"Chart": artifact},
        used_titles=set(),
        summary_sheet=summary_sheet,
    )

    rows = list(workbook["Chart"].iter_rows(values_only=True))
    assert rows[1:4] == [("series", 10.0, "a"), ("series", 20.0, "b"), ("series", 30.0, "c")]


def test_build_board_export_requires_artifacts(tmp_path) -> None:
    try:
        build_board_export(
            title="Test",
            artifacts=[],
            output_dir=tmp_path,
            export_format="docx",
        )
    except ValueError as exc:
        assert "артефакт" in str(exc).lower()
    else:
        raise AssertionError("expected ValueError for empty artifacts")


def test_xlsx_table_export_prefers_raw_numeric_values() -> None:
    artifact = {
        "data": {
            "format": "split",
            "data": {"columns": ["План"], "data": [["1,4 млн ₽"]]},
            "export_data": {"columns": ["План"], "data": [[1437258.4810761]]},
        }
    }

    assert _split_table_payload(artifact) == (["План"], [[1437258.4810761]])


def test_planfact_xlsx_export_adds_filterable_validation_sheets(tmp_path) -> None:
    from openpyxl import load_workbook

    artifact = {
        "id": "result",
        "type": "table",
        "text": "Результат",
        "data": {
            "format": "split",
            "data": {"columns": ["ЦФО", "Среднее"], "data": [["A", 10.0]]},
        },
    }
    validation_tables = {
        "Контроль": {
            "columns": [
                "Строка результата",
                "ЦФО",
                "Показатель",
                "Значение результата",
                "Проверка Excel",
                "Разница",
                "Статус",
            ],
            "rows": [
                [
                    1,
                    "A",
                    "Среднее",
                    10.0,
                    "=AVERAGEIF('Расчетная детализация'!A:A,A2,C:C)",
                    '=IF(OR(E2="",D2=""),"",E2-D2)',
                    '=IF(F2="","НЕ ПРОВЕРЕНО",IF(ABS(F2)<=0.01,"OK","РАСХОЖДЕНИЕ"))',
                ]
            ],
        },
        "Расчетная детализация": {
            "columns": ["Строка результата", "ЦФО", "Отклонение"],
            "rows": [
                [1, "A", 10.0],
                [1, "A", 10.0],
            ],
        },
        "Первичка факт": {
            "columns": ["Строка данных", "Дата", "Сумма"],
            "rows": [[1, "2026-01-10", 110.0]],
        },
    }

    result = build_board_export(
        title="Проверка",
        artifacts=[artifact],
        output_dir=tmp_path,
        export_format="xlsx",
        planfact_validation_tables=validation_tables,
    )

    workbook = load_workbook(result.file_path, data_only=False)
    assert "Как проверить" in workbook.sheetnames
    assert "Контроль" in workbook.sheetnames
    assert "Расчетная детализация" in workbook.sheetnames
    assert "Первичка факт" in workbook.sheetnames

    validation = workbook["Контроль"]
    assert validation.freeze_panes == "A2"
    assert validation.auto_filter.ref == "A1:G2"
    assert validation["A1"].value == "Строка результата"
    assert validation["E2"].value.startswith("=AVERAGEIF")
    assert validation["G2"].value == ('=IF(F2="","НЕ ПРОВЕРЕНО",IF(ABS(F2)<=0.01,"OK","РАСХОЖДЕНИЕ"))')

    source = workbook["Первичка факт"]
    assert source.auto_filter.ref == "A1:C2"
    assert source.freeze_panes == "A2"
